from flask import Flask, render_template, request, jsonify, send_file
import openpyxl
from datetime import datetime, date, time as dtime
import os, json, tempfile, io
import psycopg2, psycopg2.extras

app = Flask(__name__)

DATABASE_URL = os.environ.get('DATABASE_URL', '')
EXCEL_NAME   = 'CONSOLIDADO INFORMES CARGUES DE POLLO 2026.xlsx'
DATA_DIR     = os.environ.get('DATA_DIR', os.path.join(os.path.dirname(__file__), 'data'))
CONFIG_FILE  = os.path.join(DATA_DIR, 'config.json')
SHEET_NAME   = 'CONSOLIDADO 2026'
DATA_START   = 6
os.makedirs(DATA_DIR, exist_ok=True)

# ── DB ────────────────────────────────────────────────────────────────────────
def get_conn():
    url = DATABASE_URL
    if url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)
    return psycopg2.connect(url)

def init_db():
    if not DATABASE_URL: return
    import time
    for attempt in range(10):
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute('''CREATE TABLE IF NOT EXISTS entries (
                        id SERIAL PRIMARY KEY,
                        dia INTEGER, fecha DATE NOT NULL, viaje INTEGER,
                        granja TEXT, cuadrilla TEXT, personas INTEGER,
                        conductor TEXT, placa TEXT,
                        hora_inicio TIME, hora_salida TIME,
                        cantidad INTEGER, total_dia INTEGER
                    )''')
                conn.commit()
            return
        except Exception as e:
            if attempt < 9:
                time.sleep(3)
            else:
                print(f'DB init failed after 10 attempts: {e}')

init_db()

# ── Config ────────────────────────────────────────────────────────────────────
def load_config():
    cfg = {}
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f: cfg = json.load(f)
    if os.environ.get('GROQ_API_KEY'):
        cfg['api_key'] = os.environ['GROQ_API_KEY']
        cfg['proveedor'] = 'groq'
    return cfg

def save_config(cfg):
    with open(CONFIG_FILE, 'w') as f: json.dump(cfg, f)

# ── Helpers ───────────────────────────────────────────────────────────────────
def fmt_time(t):
    if t is None: return ''
    if isinstance(t, dtime): return t.strftime('%H:%M')
    if isinstance(t, datetime): return t.strftime('%H:%M')
    try: return f'{t.hour:02d}:{t.minute:02d}'
    except: return str(t)

def parse_time(s):
    if not s: return None
    for fmt in ('%H:%M', '%H:%M:%S'):
        try:
            dt = datetime.strptime(s.strip(), fmt)
            return dtime(dt.hour, dt.minute, dt.second)
        except: pass
    return None

def to_time(val):
    if val is None: return None
    if isinstance(val, dtime): return val
    if isinstance(val, datetime): return val.time()
    if isinstance(val, str): return parse_time(val)
    return None

def to_int(val):
    if val is None: return None
    if isinstance(val, (int, float)): return int(val)
    if isinstance(val, str):
        try: return int(float(val.replace(',','.')))
        except: return None
    return None

# ── DB Operations ─────────────────────────────────────────────────────────────
def get_entries():
    if not DATABASE_URL: return []
    with get_conn() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
            cur.execute('SELECT * FROM entries ORDER BY fecha, viaje, id')
            rows = cur.fetchall()
    result = []
    for r in rows:
        result.append({
            'dia': r['dia'],
            'fecha': r['fecha'].strftime('%Y-%m-%d') if r['fecha'] else '',
            'viaje': r['viaje'],
            'granja': r['granja'] or '',
            'cuadrilla': r['cuadrilla'] or '',
            'personas': r['personas'],
            'conductor': r['conductor'] or '',
            'placa': r['placa'] or '',
            'hora_inicio': fmt_time(r['hora_inicio']),
            'hora_salida': fmt_time(r['hora_salida']),
            'cantidad': r['cantidad'],
            'total_dia': r['total_dia'],
        })
    return result

def next_viaje(fecha_s):
    if not DATABASE_URL: return 1
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute('SELECT MAX(viaje) FROM entries WHERE fecha = %s', (fecha_s,))
            row = cur.fetchone()
    v = row[0] if row and row[0] else 0
    return v + 1

def last_dia():
    if not DATABASE_URL: return 1
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute('SELECT MAX(dia) FROM entries')
            row = cur.fetchone()
    return row[0] if row and row[0] else 1

def insert_entry(fecha_s, viaje, dia_num, data):
    fecha_dt = datetime.strptime(fecha_s, '%Y-%m-%d').date()
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute('''INSERT INTO entries
                (dia, fecha, viaje, granja, cuadrilla, personas, conductor, placa, hora_inicio, hora_salida, cantidad, total_dia)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
                (dia_num, fecha_dt, viaje,
                 (data.get('granja','') or '').upper() or None,
                 (data.get('cuadrilla','CARGUEROS') or 'CARGUEROS').upper(),
                 int(data.get('personas',11) or 11),
                 (data.get('conductor','') or '').upper() or None,
                 (data.get('placa','') or '').upper() or None,
                 parse_time(data.get('hora_inicio','')),
                 parse_time(data.get('hora_salida','')),
                 int(data['cantidad']) if data.get('cantidad') else None,
                 int(data['total_dia']) if data.get('total_dia') else None))
        conn.commit()

def generate_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.cell(row=3, column=5).value = 'REPORTE DIARIO DE CARGUE  SAVICOL'
    headers = [None, None, 'FECHA ', 'Nº VIAJES', 'GRANJA ', 'CUADRILLA',
               'CANTIDAD DE PERSONAS ', 'NOMBRE CONDUCTOR', 'PLACA',
               'HORA INICIO', 'HORA SALIDA ', 'CANTIDAD', 'TOTAL POLLOS DIA ']
    for col, h in enumerate(headers, 1):
        ws.cell(row=DATA_START-1, column=col).value = h
    for i, e in enumerate(get_entries()):
        r = DATA_START + i
        ws.cell(r, 2).value = e['dia']
        ws.cell(r, 3).value = datetime.strptime(e['fecha'], '%Y-%m-%d') if e['fecha'] else None
        ws.cell(r, 4).value = e['viaje']
        ws.cell(r, 5).value = e['granja']
        ws.cell(r, 6).value = e['cuadrilla']
        ws.cell(r, 7).value = e['personas']
        ws.cell(r, 8).value = e['conductor']
        ws.cell(r, 9).value = e['placa']
        ws.cell(r, 10).value = parse_time(e['hora_inicio'])
        ws.cell(r, 11).value = parse_time(e['hora_salida'])
        ws.cell(r, 12).value = e['cantidad']
        ws.cell(r, 13).value = e['total_dia']
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Prompt ────────────────────────────────────────────────────────────────────
PROMPT = """Eres un asistente que extrae datos de reportes de cargue de pollos en Colombia.

REGLAS:
1. PLACAS colombianas: exactamente 3 letras + 3 números (ej: TLO881, EQY779).
   Whisper las transcribe mal: "EQY es 181"→EQY181, "EQY setenta y ocho uno"→EQY781.
2. HORAS: "once treinta y cinco"=11:35, "tres y veintidós"=03:22. Formato HH:MM 24h.
3. NÚMEROS: "tres mil quinientos sesenta y cuatro"=3564.
4. NOMBRES: nombres propios colombianos.

Para cada viaje devuelve:
- conductor: nombre en MAYÚSCULAS
- placa: 3 letras + 3 números, sin espacios
- hora_inicio: HH:MM 24h
- hora_salida: HH:MM 24h
- cantidad: entero
- total_dia: entero si se menciona como total del día, sino null
- granja: MALAGUEÑA/GARCERAS/SAN JOSE/HUERTAS/etc. o null
- cuadrilla: "CARGUEROS" por defecto
- personas: 11 por defecto

Responde SOLO con JSON válido:
{"viajes": [...]}

Transcripción:
"""

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html', today=date.today().strftime('%Y-%m-%d'))

@app.route('/api/config', methods=['GET','POST'])
def api_config():
    if request.method == 'POST':
        cfg = load_config(); cfg.update(request.json); save_config(cfg)
        return jsonify({'ok': True})
    cfg = load_config(); cfg.pop('api_key', None)
    return jsonify(cfg)

@app.route('/api/entries')
def api_entries():
    return jsonify(get_entries())

@app.route('/api/next-viaje')
def api_next_viaje():
    fecha = request.args.get('fecha', date.today().strftime('%Y-%m-%d'))
    return jsonify({'next_viaje': next_viaje(fecha)})

@app.route('/api/has-key')
def api_has_key():
    cfg = load_config()
    return jsonify({'has_key': bool(cfg.get('api_key','').strip())})

@app.route('/api/db-status')
def api_db_status():
    if not DATABASE_URL:
        return jsonify({'ok': False, 'error': 'Base de datos no configurada'})
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT COUNT(*) FROM entries')
                count = cur.fetchone()[0]
        return jsonify({'ok': True, 'count': count})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@app.route('/api/transcribe', methods=['POST'])
def api_transcribe():
    if 'audio' not in request.files:
        return jsonify({'ok':False,'error':'No se recibió archivo'}), 400
    f = request.files['audio']
    api_key = load_config().get('api_key','')
    if not api_key:
        return jsonify({'ok':False,'error':'Configura la clave API primero'}), 400
    suffix = os.path.splitext(f.filename)[1] or '.ogg'
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        f.save(tmp.name); tmp_path = tmp.name
    try:
        from groq import Groq
        client = Groq(api_key=api_key)
        with open(tmp_path, 'rb') as audio_file:
            result = client.audio.transcriptions.create(
                file=(f.filename, audio_file),
                model='whisper-large-v3',
                language='es',
                response_format='text'
            )
        texto = result if isinstance(result, str) else result.text
        return jsonify({'ok':True,'texto':texto.strip()})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500
    finally:
        os.unlink(tmp_path)

@app.route('/api/parse', methods=['POST'])
def api_parse():
    try:
        body = request.json
        texto = body.get('texto','').strip()
        body_key = body.get('api_key','')
        api_key = load_config().get('api_key','') if (not body_key or body_key=='__server__') else body_key
        fecha = body.get('fecha', date.today().strftime('%Y-%m-%d'))
        proveedor = load_config().get('proveedor','groq')
        if not api_key: return jsonify({'ok':False,'error':'Clave API no configurada'}), 400
        if not texto:   return jsonify({'ok':False,'error':'Texto vacío'}), 400
        full = PROMPT + texto
        if proveedor == 'anthropic':
            import anthropic
            c = anthropic.Anthropic(api_key=api_key)
            msg = c.messages.create(model='claude-haiku-4-5-20251001', max_tokens=2048,
                                    messages=[{'role':'user','content':full}])
            raw = msg.content[0].text.strip()
        else:
            from groq import Groq
            c = Groq(api_key=api_key)
            chat = c.chat.completions.create(model='llama-3.3-70b-versatile',
                messages=[{'role':'user','content':full}], max_tokens=2048, temperature=0.1)
            raw = chat.choices[0].message.content.strip()
        if raw.startswith('```'): raw = raw.split('\n',1)[1].rsplit('```',1)[0].strip()
        viajes = json.loads(raw).get('viajes',[])
        v = next_viaje(fecha)
        for i, x in enumerate(viajes): x['viaje_num'] = v + i
        return jsonify({'ok':True,'viajes':viajes})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

@app.route('/api/add-bulk', methods=['POST'])
def api_add_bulk():
    body = request.json
    viajes = body.get('viajes',[])
    fecha_s = body.get('fecha', date.today().strftime('%Y-%m-%d'))
    if not viajes: return jsonify({'ok':False,'error':'Sin viajes'}), 400
    try:
        entries = get_entries()
        fe = [e for e in entries if e['fecha'] == fecha_s]
        dia_num = fe[0]['dia'] if fe and fe[0]['dia'] else last_dia()
        vs = next_viaje(fecha_s)
        for i, v in enumerate(viajes):
            insert_entry(fecha_s, vs+i, dia_num, v)
        return jsonify({'ok':True,'guardados':len(viajes)})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

@app.route('/api/add', methods=['POST'])
def api_add():
    data = request.json
    try:
        fecha_s = data.get('fecha', date.today().strftime('%Y-%m-%d'))
        viaje = data.get('viaje') or next_viaje(fecha_s)
        entries = get_entries()
        fe = [e for e in entries if e['fecha'] == fecha_s]
        dia_num = fe[0]['dia'] if fe and fe[0]['dia'] else last_dia()
        insert_entry(fecha_s, viaje, dia_num, data)
        return jsonify({'ok':True,'viaje':viaje})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

@app.route('/api/delete-last', methods=['POST'])
def api_delete_last():
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute('SELECT id FROM entries ORDER BY id DESC LIMIT 1')
                row = cur.fetchone()
                if not row:
                    return jsonify({'ok':False,'error':'No hay registros'})
                cur.execute('DELETE FROM entries WHERE id = %s', (row[0],))
            conn.commit()
        return jsonify({'ok':True})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

@app.route('/api/download-excel')
def api_download_excel():
    try:
        buf = generate_excel()
        return send_file(buf, as_attachment=True, download_name=EXCEL_NAME,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/api/upload-excel', methods=['POST'])
def api_upload_excel():
    if 'excel' not in request.files:
        return jsonify({'ok':False,'error':'No se recibió archivo'}), 400
    f = request.files['excel']
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb[SHEET_NAME]
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute('DELETE FROM entries')
                count = 0
                for r in ws.iter_rows(min_row=DATA_START, values_only=True):
                    if all(v is None for v in r): continue
                    fecha = r[2]
                    if fecha is None: continue
                    fecha_dt = fecha.date() if isinstance(fecha, datetime) else fecha
                    cur.execute('''INSERT INTO entries
                        (dia, fecha, viaje, granja, cuadrilla, personas, conductor, placa, hora_inicio, hora_salida, cantidad, total_dia)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
                        (r[1], fecha_dt, r[3],
                         str(r[4] or '').upper() or None,
                         str(r[5] or '').upper() or None,
                         int(r[6]) if r[6] else None,
                         str(r[7] or '').upper() or None,
                         str(r[8] or '').upper() or None,
                         to_time(r[9]), to_time(r[10]),
                         to_int(r[11]), to_int(r[12])))
                    count += 1
            conn.commit()
        return jsonify({'ok':True,'registros':count})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

@app.route('/api/open-folder', methods=['POST'])
def api_open_folder():
    import subprocess
    subprocess.Popen(['open', DATA_DIR])
    return jsonify({'ok':True})

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT',5050)))
